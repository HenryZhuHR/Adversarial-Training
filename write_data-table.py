"""

"""


import os
import tqdm
import openpyxl
import cv2
import numpy
from PIL import Image
import torch
from torch import nn
from torch import Tensor
from torchvision import transforms

from api_robustModel import RobustResnet34


DEVICE = 'cpu'
DATASET_BASEDIR = 'E:/datasets/gc10_dsets'
# test224
TRANSFORM = transforms.Compose([transforms.Resize(224), transforms.ToTensor()])

if __name__ == '__main__':
    model_path = 'api_robustModel/models/resnet34-adv-8.pt'
    model = RobustResnet34(model_weight_path=model_path, device=DEVICE)

    # workbook
    wb = openpyxl.load_workbook('./data.xlsx')

    # workspace
    print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print(ws)

    print()
    index = 0
    pbar = tqdm.tqdm(ws.iter_rows(min_row=3))
    cls_n = ''
    for data in pbar:
        # print(data)
        # /airplane/airplane_0003.jpg  --> str.split --> ['','airplane','airplane_0003.jpg']
        _, class_name, file_name = str(data[0].value).split('/')
        if class_name != cls_n:
            index = 0
            cls_n = class_name

        # print(class_name,file_name)
        images_file_list = os.listdir(os.path.join(
            DATASET_BASEDIR, 'pgd_8_255_demo', class_name))
        image_path = os.path.join(
            DATASET_BASEDIR, 'pgd_8_255_demo', class_name, images_file_list[index])
        image: numpy.ndarray = cv2.imread(image_path)

        if image.shape[2] != 3:
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
        else:
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image_tensor: Tensor = TRANSFORM(Image.fromarray(image))

        prob_list, index_list, predict_labels = model.top_k(image_tensor, k=5)
        prob = prob_list[0]

        # robust
        data[8].value = 1 if class_name == predict_labels[0] else 0

        # after rebuilt
        images_file_list = os.listdir(os.path.join(
            DATASET_BASEDIR, 'recon_pgd_8_255_demo', class_name))
        image_path = os.path.join(
            DATASET_BASEDIR, 'recon_pgd_8_255_demo', class_name, images_file_list[index])
        image: numpy.ndarray = cv2.imread(image_path)
        if image.shape[2] != 3:
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
        else:
            image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image_tensor: Tensor = TRANSFORM(Image.fromarray(image))

        prob_list, index_list, predict_labels = model.top_k(image_tensor, k=5)
        data[13].value = 1 if class_name == predict_labels[0] else 0
        rebuild_prob = prob_list[0]

        increase = (rebuild_prob-prob)*100
        if increase >= 0:
            if increase>=10:
                increase_str = '%2d%%↑↑' % increase
            else:
                increase_str = '%2d%%↑ ' % increase
        else:
            increase_str = '     ' #'%2d%%↓ ' % (-increase)
        pbar.set_description('%3d  %s: %s ' % (index, class_name, file_name) +
                             ' prob: %.4f -> %.4f ' % (prob, rebuild_prob) +
                             increase_str)
        index += 1

    wb.save('./data-res-8.xlsx')
